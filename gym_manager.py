"""
Database-backed Gym Manager
Uses PostgreSQL via SQLAlchemy instead of JSON files
Build Tag: 20251218-2135
"""

from models import (Base, Member, Fee, Attendance, Expense, Gym, User, 
                    get_database_url, get_session, MemberNote, BodyMeasurement)
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from sqlalchemy import func, extract
from sqlalchemy.exc import IntegrityError
import os
import json

class GymManager:
    def __init__(self, user_email):
        """Initialize with user's email"""
        self.user_email = user_email
        self.session = get_session()
        
        if not self.session:
            print(f"⚠️ GymManager: DB Connection FAILED for {user_email}. Falling back to JSON.")
            self.legacy = True
            self.data_file = f"gym_data/{user_email}.json"
            self.data = self.load_legacy_data()
            return

        self.legacy = False
        # Get or create user's gym
        user = self.session.query(User).filter_by(email=user_email).first()
        if user:
            self.gym = self.session.query(Gym).filter_by(user_id=user.id).first()
            if not self.gym:
                # Create default gym for user
                self.gym = Gym(
                    user_id=user.id,
                    name='Gym Manager',
                    currency='Rs'
                )
                self.session.add(self.gym)
                self.session.commit()
        else:
            self.gym = None

    def load_legacy_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    return json.load(f)
            except:
                pass
        return {
            'members': {},
            'fees': {},
            'expenses': [],
            'attendance': {},
            'gym_details': {'name': 'Gym Manager', 'logo': None, 'currency': 'Rs'}
        }

    def get_gym_details(self) -> Dict:
        """Get gym name, logo, and currency"""
        if self.legacy:
            return self.data.get('gym_details', {'name': 'Gym Manager', 'logo': None, 'currency': 'Rs'})
        
        if not self.gym:
            return {'name': 'Gym Manager', 'logo': None, 'currency': 'Rs'}
        
        return {
            'name': self.gym.name,
            'logo': self.gym.logo_url,
            'currency': self.gym.currency
        }

    def update_gym_details(self, name: str, logo_path: Optional[str] = None, currency: str = 'Rs') -> bool:
        """Update gym name, logo, and currency"""
        if self.legacy:
            self.data['gym_details'] = {
                'name': name,
                'logo': logo_path or self.data.get('gym_details', {}).get('logo'),
                'currency': currency
            }
            self.save_legacy_data()
            return True

        if not self.gym:
            return False
        
        self.gym.name = name
        self.gym.currency = currency
        if logo_path:
            self.gym.logo_url = logo_path
            
        self.session.commit()
        return True

    def add_member(self, name, phone, photo=None, membership_type='Gym', joined_date=None, is_trial=False, email=None):
        """Add a new member"""
        if self.legacy:
            # Simple ID generation for legacy
            new_id = str(int(max(self.data['members'].keys(), default=0)) + 1)
            self.data['members'][new_id] = {
                'id': new_id,
                'name': name,
                'phone': phone,
                'photo': photo,
                'membership_type': membership_type,
                'joined_date': joined_date if joined_date else datetime.now().strftime('%Y-%m-%d'),
                'is_trial': is_trial,
                'active': True,
                'email': email
            }
            self.save_legacy_data()
            return new_id

        if not self.gym:
            return None
        
        if not joined_date:
            joined_date = datetime.now().date()
        elif isinstance(joined_date, str):
            joined_date = datetime.strptime(joined_date, '%Y-%m-%d').date()
            
        trial_end = None
        if is_trial:
            trial_end = joined_date + timedelta(days=3)
            
        member = Member(
            gym_id=self.gym.id,
            name=name,
            phone=phone,
            email=email,
            photo_url=photo,
            joined_date=joined_date,
            membership_type=membership_type,
            is_trial=is_trial,
            trial_end_date=trial_end
        )
        
        self.session.add(member)
        self.session.commit()
        return member.id

    def update_member(self, member_id, name, phone, membership_type, joined_date=None, email=None):
        """Update member details in database"""
        member = self.session.query(Member).get(int(member_id))
        if not member:
            return False
        
        member.name = name
        member.phone = phone
        member.membership_type = membership_type
        member.email = email
        
        if joined_date:
            if isinstance(joined_date, str):
                member.joined_date = datetime.strptime(joined_date, '%Y-%m-%d').date()
            else:
                member.joined_date = joined_date
        
        self.session.commit()
        return True

    def delete_member(self, member_id):
        """Delete member and their records"""
        if self.legacy:
            if str(member_id) in self.data['members']:
                del self.data['members'][str(member_id)]
                if str(member_id) in self.data['fees']:
                    del self.data['fees'][str(member_id)]
                if str(member_id) in self.data['attendance']:
                    del self.data['attendance'][str(member_id)]
                self.save_legacy_data()
                return True
            return False

        member = self.session.query(Member).get(int(member_id))
        if member:
            self.session.delete(member)
            self.session.commit()
            return True
        return False

    def get_member(self, member_id):
        """Get member by ID"""
        if self.legacy:
            return self.data['members'].get(str(member_id))

        member = self.session.query(Member).get(int(member_id))
        if not member:
            return None
        return self._member_to_dict(member)

    def get_all_members(self):
        """Get all members"""
        if self.legacy:
            return list(self.data['members'].values())

        if not self.gym:
            return []
        members = self.session.query(Member).filter_by(gym_id=self.gym.id, is_active=True).all()
        return [self._member_to_dict(m) for m in members]

    def save_legacy_data(self):
        """Save legacy data to JSON"""
        if self.legacy:
            with open(self.data_file, 'w') as f:
                json.dump(self.data, f, indent=2)

    def reset_data(self):
        """Reset all data for this gym"""
        if self.legacy:
            self.data = {'members': {}, 'fees': {}, 'attendance': {}, 'expenses': []}
            self.save_legacy_data()
            return True
        
        if not self.gym:
            return False
            
        try:
            # Delete all attendance records for this gym's members
            member_ids = [m.id for m in self.session.query(Member).filter_by(gym_id=self.gym.id).all()]
            if member_ids:
                self.session.query(Attendance).filter(Attendance.member_id.in_(member_ids)).delete(synchronize_session=False)
                self.session.query(Fee).filter(Fee.member_id.in_(member_ids)).delete(synchronize_session=False)
            
            # Delete all members for this gym
            self.session.query(Member).filter_by(gym_id=self.gym.id).delete()
            
            # Delete all expenses for this gym
            self.session.query(Expense).filter_by(gym_id=self.gym.id).delete()
            
            self.session.commit()
            return True
        except Exception as e:
            self.session.rollback()
            print(f"Reset error: {e}")
            return False

    def _member_to_dict(self, member):
        """Convert SQLAlchemy member object to dictionary"""
        return {
            'id': str(member.id),
            'name': member.name,
            'phone': member.phone,
            'email': member.email,
            'photo': member.photo_url,
            'joined_date': member.joined_date.strftime('%Y-%m-%d'),
            'membership_type': member.membership_type,
            'is_trial': member.is_trial,
            'trial_end_date': member.trial_end_date.strftime('%Y-%m-%d') if member.trial_end_date else None,
            'active': member.is_active
        }

    def record_fee(self, member_id, month, amount, date=None):
        """Record a fee payment"""
        if self.legacy:
            if str(member_id) not in self.data['fees']:
                self.data['fees'][str(member_id)] = {}
            
            if not date:
                date = datetime.now()
            elif isinstance(date, str):
                try:
                    date = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                except:
                    date = datetime.strptime(date, '%Y-%m-%d')
            
            self.data['fees'][str(member_id)][month] = {
                'amount': amount,
                'date': date.strftime('%Y-%m-%d %H:%M:%S')
            }
            self.save_legacy_data()
            return True

        if not date:
            date = datetime.now()
        elif isinstance(date, str):
            try:
                date = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
            except:
                date = datetime.strptime(date, '%Y-%m-%d')
                
        # Check if already paid
        existing = self.session.query(Fee).filter_by(member_id=int(member_id), month=month).first()
        if existing:
            return False
            
        fee = Fee(
            member_id=int(member_id),
            month=month,
            amount=amount,
            paid_date=date
        )
        self.session.add(fee)
        self.session.commit()
        return True

    def get_member_fees(self, member_id):
        """Get all fee records for a member"""
        if self.legacy:
            member_fees = self.data['fees'].get(str(member_id), {})
            return [{
                'month': m,
                'amount': float(info.get('amount', 0)),
                'paid_date': info.get('date', info.get('timestamp', 'N/A')),
                'notes': info.get('notes', '')
            } for m, info in member_fees.items()]

        fees = self.session.query(Fee).filter_by(member_id=int(member_id)).order_by(Fee.month.desc()).all()
        return [{
            'month': f.month,
            'amount': float(f.amount),
            'paid_date': f.paid_date.strftime('%Y-%m-%d %H:%M:%S'),
            'notes': ''  # Notes not supported yet in DB
        } for f in fees]

    def is_fee_paid(self, member_id, month):
        """Check if fee is paid for a specific month"""
        fee = self.session.query(Fee).filter_by(member_id=int(member_id), month=month).first()
        return fee is not None
    
    def pay_fee(self, member_id, month, amount, payment_date=None, notes=None):
        """Alias for record_fee - notes parameter ignored for now"""
        return self.record_fee(member_id, month, amount, payment_date)

    def get_payment_status(self, month=None):
        """Get paid/unpaid members for a month"""
        if not month:
            month = datetime.now().strftime('%Y-%m')
            
        if self.legacy:
            paid = []
            unpaid = []
            fees = self.data.get('fees', {})
            members = self.data.get('members', {})
            
            for mid, m in members.items():
                member_data = m.copy()
                if mid in fees and month in fees[mid]:
                    fee_info = fees[mid][month]
                    member_data['amount'] = fee_info.get('amount', 0)
                    # Safety check for 'date' key
                    member_data['date'] = fee_info.get('date', fee_info.get('timestamp', 'N/A'))
                    paid.append(member_data)
                else:
                    unpaid.append(member_data)
            return {'paid': paid, 'unpaid': unpaid}

        if not self.gym:
            return {'paid': [], 'unpaid': []}
            
        all_members = self.session.query(Member).filter_by(gym_id=self.gym.id, is_active=True).all()
        paid_records = self.session.query(Fee).filter(Fee.member_id.in_([m.id for m in all_members]), Fee.month == month).all()
        
        paid_ids = {f.member_id: f for f in paid_records}
        
        paid = []
        unpaid = []
        
        for m in all_members:
            m_dict = self._member_to_dict(m)
            if m.id in paid_ids:
                f = paid_ids[m.id]
                m_dict['amount'] = float(f.amount)
                m_dict['date'] = f.paid_date.strftime('%Y-%m-%d %H:%M:%S')
                paid.append(m_dict)
            else:
                unpaid.append(m_dict)
                
        return {'paid': paid, 'unpaid': unpaid}

    def get_revenue(self, month=None):
        """Get total revenue for a month"""
        if self.legacy:
            total = 0.0
            for mid, member_fees in self.data.get('fees', {}).items():
                if month:
                    if month in member_fees:
                        total += float(member_fees[month].get('amount', 0))
                else:
                    for m, info in member_fees.items():
                        total += float(info.get('amount', 0))
            return total

        if not self.gym:
            return 0.0
            
        query = self.session.query(func.sum(Fee.amount))
        if month:
            query = query.filter(Fee.month == month)
            
        # Ensure only for current gym's members
        query = query.join(Member).filter(Member.gym_id == self.gym.id)
        
        result = query.scalar()
        return float(result) if result else 0.0

    def add_expense(self, category, amount, date, description=''):
        """Add an expense record"""
        if self.legacy:
            if isinstance(date, str):
                date_str = date
            else:
                date_str = date.strftime('%Y-%m-%d')
                
            self.data['expenses'].append({
                'category': category,
                'amount': amount,
                'date': date_str,
                'description': description
            })
            self.save_legacy_data()
            return True

        if not self.gym:
            return False
            
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d').date()
            
        expense = Expense(
            gym_id=self.gym.id,
            category=category,
            amount=amount,
            date=date,
            description=description
        )
        self.session.add(expense)
        self.session.commit()
        return True

    def get_expenses(self, month=None):
        """Get expenses history"""
        if self.legacy:
            expenses = self.data.get('expenses', [])
            if month:
                expenses = [e for e in expenses if e.get('date', '').startswith(month)]
            # Sort by date desc
            return sorted(expenses, key=lambda x: x.get('date', ''), reverse=True)

        if not self.gym:
            return []
            
        query = self.session.query(Expense).filter_by(gym_id=self.gym.id)
        if month:
            # Check database dialect
            if self.session.bind.dialect.name == 'postgresql':
                query = query.filter(func.to_char(Expense.date, 'YYYY-MM') == month)
            else:
                # SQLite fallback
                query = query.filter(func.strftime('%Y-%m', Expense.date) == month)
            
        expenses = query.order_by(Expense.date.desc()).all()
        return [{
            'id': e.id,
            'category': e.category,
            'amount': float(e.amount),
            'date': e.date.strftime('%Y-%m-%d'),
            'description': e.description
        } for e in expenses]
    
    def calculate_profit_loss(self, month=None):
        """Calculate profit/loss for a month"""
        if not month:
            month = datetime.now().strftime('%Y-%m')
        
        # Get revenue (fees collected)
        revenue = self.get_revenue(month)
        
        # Get expenses
        expenses_list = self.get_expenses(month)
        total_expenses = sum(e.get('amount', 0) for e in expenses_list)
        
        # Calculate profit
        net_profit = revenue - total_expenses
        
        # Calculate profit margin
        profit_margin = round((net_profit / revenue * 100), 2) if revenue > 0 else 0
        
        return {
            'revenue': revenue,
            'expenses': total_expenses,
            'net_profit': net_profit,
            'profit_margin': profit_margin,
            'month': month
        }

    def log_attendance(self, member_id, emotion=None, confidence=None):
        """Log member check-in"""
        if self.legacy:
            if str(member_id) not in self.data['attendance']:
                self.data['attendance'][str(member_id)] = []
            
            self.data['attendance'][str(member_id)].append({
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'emotion': emotion,
                'confidence': confidence
            })
            self.save_legacy_data()
            return True

        attendance = Attendance(
            member_id=int(member_id),
            # check_in_time removed - using default created_at
            emotion=emotion,
            confidence=confidence
        )
        self.session.add(attendance)
        self.session.commit()
        return True

    def get_attendance(self, member_id):
        """Get attendance history for a member"""
        if self.legacy:
            return sorted(self.data['attendance'].get(str(member_id), []), 
                          key=lambda x: x.get('timestamp', ''), reverse=True)

        records = self.session.query(Attendance).filter_by(member_id=int(member_id)).order_by(Attendance.created_at.desc()).all()
        return [{
            'timestamp': r.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'emotion': r.emotion,
            'confidence': float(r.confidence) if r.confidence else None
        } for r in records]

    def bulk_import_members(self, filepath):
        """Import members from Excel/CSV with batch processing - NO PANDAS"""
        import csv
        from openpyxl import load_workbook
        
        try:
            # Read file based on extension
            rows_data = []
            headers = []
            
            if filepath.endswith('.csv'):
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames
                    rows_data = list(reader)
            else:
                # Excel file - use openpyxl instead of pandas
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    rows_data.append(row_dict)
                wb.close()
            
            success = 0
            errors = []
            new_members = []
            fee_records = []  # NEW: Store fee records for paid months
            
            # Legacy Mode: Fallback to single add
            if self.legacy:
                for idx, row in enumerate(rows_data):
                    try:
                        name = str(row['Name']).strip() if row.get('Name') else ''
                        phone = str(row['Phone']).strip() if row.get('Phone') else ''
                        self.add_member(name, phone)
                        success += 1
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                return success, len(errors), errors

            # Database Mode: Batch Processing
            # 1. Get existing phones to avoid duplicates
            existing_phones = {m.phone for m in self.session.query(Member).filter_by(gym_id=self.gym.id).all()}
            
            for index, row in enumerate(rows_data):
                try:
                    name = str(row['Name']).strip() if row.get('Name') else ''
                    phone = str(row['Phone']).strip() if row.get('Phone') else ''
                    
                    if not name or not phone:
                        errors.append(f"Row {index}: Missing name or phone")
                        continue
                    
                    if phone in existing_phones:
                        errors.append(f"Row {index}: Member with phone {phone} already exists")
                        continue
                        
                    # Handle optional fields
                    email = str(row['Email']).strip() if row.get('Email') and row['Email'] else None
                    membership_type = str(row['Membership Type']).strip() if row.get('Membership Type') and row['Membership Type'] else 'Gym'
                    
                    joined_date = datetime.now().date()
                    if row.get('Joined Date') and row['Joined Date']:
                        try:
                            jd = row['Joined Date']
                            if isinstance(jd, str):
                                joined_date = datetime.strptime(jd, '%Y-%m-%d').date()
                            elif hasattr(jd, 'date'):
                                joined_date = jd.date() if callable(jd.date) else jd
                        except:
                            pass

                    member = Member(
                        gym_id=self.gym.id,
                        name=name,
                        phone=phone,
                        email=email,
                        membership_type=membership_type,
                        joined_date=joined_date,
                        photo_url=None,
                        is_trial=False,
                        is_active=True  # Force active status
                    )
                    new_members.append(member)
                    existing_phones.add(phone) # Prevent duplicates within same file
                    
                    # NEW: Check for Paid Month column to auto-create fee record
                    paid_month = row.get('Paid Month') or row.get('paid_month')
                    amount = row.get('Amount') or row.get('amount')
                    
                    if paid_month and str(paid_month).strip():
                        # Store for later processing after member is added
                        fee_records.append({
                            'phone': phone,
                            'month': str(paid_month).strip(),
                            'amount': float(amount) if amount else 0.0
                        })
                    
                    success += 1
                    
                except Exception as e:
                    errors.append(f"Row {index}: {str(e)}")
            
            if new_members:
                try:
                    # Batch processing to prevent timeout
                    total_to_add = len(new_members)
                    batch_size = 10
                    
                    for i in range(0, total_to_add, batch_size):
                        batch = new_members[i:i + batch_size]
                        self.session.add_all(batch)
                        self.session.flush()
                        self.session.commit()
                        print(f"✓ Bulk Import: Committed batch {i}-{min(i+batch_size, total_to_add)}")
                    
                    # NEW: Process fee records for paid months
                    if fee_records:
                        print(f"💰 Processing {len(fee_records)} fee records...")
                        for fee_data in fee_records:
                            try:
                                # Find member by phone
                                member = self.session.query(Member).filter_by(
                                    gym_id=self.gym.id,
                                    phone=fee_data['phone']
                                ).first()
                                
                                if member:
                                    # Create fee record
                                    fee = Fee(
                                        member_id=member.id,
                                        month=fee_data['month'],
                                        amount=fee_data['amount'],
                                        paid_date=datetime.now().date()
                                    )
                                    self.session.add(fee)
                            except Exception as e:
                                print(f"⚠️ Fee record error: {str(e)}")
                        
                        self.session.commit()
                        print(f"✅ Fee records processed successfully")
                    
                    # Final flush
                    self.session.flush()
                    
                except Exception as e:
                    self.session.rollback()
                    return success, len(errors) + 1, errors + [f"Database Commit Error: {str(e)}"]
                pass
        
            return success, len(errors), errors
            
        except Exception as e:
            return 0, 1, [str(e)]

    # ==================== MEMBER NOTES METHODS ====================
    
    def add_member_note(self, member_id, note_text):
        """Add a note to a member's profile"""
        if self.legacy:
            return False  # Notes only work in SQL mode
        
        try:
            note = MemberNote(
                member_id=int(member_id),
                note=note_text
            )
            self.session.add(note)
            self.session.commit()
            return True
        except Exception as e:
            print(f"Error adding note: {str(e)}")
            self.session.rollback()
            return False
    
    def get_member_notes(self, member_id):
        """Get all notes for a member"""
        if self.legacy:
            return []
        
        try:
            notes = self.session.query(MemberNote).filter_by(
                member_id=int(member_id)
            ).order_by(MemberNote.created_at.desc()).all()
            
            return [{
                'id': note.id,
                'note': note.note,
                'created_at': note.created_at.strftime('%Y-%m-%d %H:%M:%S')
            } for note in notes]
        except:
            return []
    
    def delete_member_note(self, note_id):
        """Delete a specific note"""
        if self.legacy:
            return False
        
        try:
            note = self.session.query(MemberNote).filter_by(id=int(note_id)).first()
            if note:
                self.session.delete(note)
                self.session.commit()
                return True
            return False
        except:
            self.session.rollback()
            return False
    
    def get_member_timeline(self, member_id):
        """Get comprehensive activity timeline for a member"""
        if self.legacy:
            return []
        
        timeline = []
        
        # Get payments
        fees = self.get_member_fees(member_id)
        for fee in fees:
            timeline.append({
                'type': 'payment',
                'icon': '💰',
                'title': f"Payment Received - {fee.get('month')}",
                'description': f"Amount: Rs {fee.get('amount')}",
                'timestamp': fee.get('paid_date'),
                'data': fee
            })
        
        # Get attendance
        attendance = self.get_attendance(member_id)
        for record in attendance:
            timeline.append({
                'type': 'checkin',
                'icon': '✅',
                'title': 'Gym Check-in',
                'description': f"Emotion: {record.get('emotion', 'N/A')}",
                'timestamp': record.get('timestamp'),
                'data': record
            })
        
        # Get notes
        notes = self.get_member_notes(member_id)
        for note in notes:
            timeline.append({
                'type': 'note',
                'icon': '📝',
                'title': 'Admin Note Added',
                'description': note.get('note')[:100] + ('...' if len(note.get('note', '')) > 100 else ''),
                'timestamp': note.get('created_at'),
                'data': note
            })
        
        # Sort by timestamp (newest first)
        timeline.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return timeline
    
    # ==================== BODY MEASUREMENTS METHODS ====================
    
    def add_body_measurement(self, member_id, weight, body_fat=None, chest=None, waist=None, arms=None, notes=None):
        """Add body measurement record for a member"""
        if self.legacy:
            return False
        
        try:
            measurement = BodyMeasurement(
                member_id=int(member_id),
                weight=float(weight) if weight else None,
                body_fat=float(body_fat) if body_fat else None,
                chest=float(chest) if chest else None,
                waist=float(waist) if waist else None,
                arms=float(arms) if arms else None,
                notes=notes
            )
            self.session.add(measurement)
            self.session.commit()
            return True
        except Exception as e:
            print(f"Error adding measurement: {str(e)}")
            self.session.rollback()
            return False
    
    def get_body_measurements(self, member_id):
        """Get all body measurements for a member"""
        if self.legacy:
            return []
        
        try:
            measurements = self.session.query(BodyMeasurement).filter_by(
                member_id=int(member_id)
            ).order_by(BodyMeasurement.recorded_at.desc()).all()
            
            return [{
                'id': m.id,
                'weight': float(m.weight) if m.weight else None,
                'body_fat': float(m.body_fat) if m.body_fat else None,
                'chest': float(m.chest) if m.chest else None,
                'waist': float(m.waist) if m.waist else None,
                'arms': float(m.arms) if m.arms else None,
                'notes': m.notes,
                'recorded_at': m.recorded_at.strftime('%Y-%m-%d')
            } for m in measurements]
        except:
            return []
    
    def delete_body_measurement(self, measurement_id):
        """Delete a body measurement record"""
        if self.legacy:
            return False
        
        try:
            measurement = self.session.query(BodyMeasurement).filter_by(id=int(measurement_id)).first()
            if measurement:
                self.session.delete(measurement)
                self.session.commit()
                return True
            return False
        except:
            self.session.rollback()
            return False
            
            
    def find_duplicates(self):
        """Find duplicate members based on name and phone"""
        if self.legacy:
            return [] # Not supported in legacy mode
            
        # Group members by name and phone
        from collections import defaultdict
        groups = defaultdict(list)
        
        all_members = self.session.query(Member).filter_by(gym_id=self.gym.id).all()
        for m in all_members:
            # Normalize key: lower case name, numeric phone
            key = (m.name.lower().strip(), "".join(filter(str.isdigit, m.phone)))
            groups[key].append(m)
            
        # Filter for actual duplicates
        duplicates = []
        for key, members in groups.items():
            if len(members) > 1:
                # Sort by joined date (keep oldest)
                members.sort(key=lambda x: x.joined_date)
                duplicates.append({
                    'primary': self._member_to_dict(members[0]),
                    'duplicates': [self._member_to_dict(m) for m in members[1:]]
                })
        return duplicates
        
    def merge_members(self):
        """Automatically merge all duplicate members"""
        if self.legacy:
            return 0
            
        duplicates = self.find_duplicates()
        merged_count = 0
        
        for group in duplicates:
            primary_id = int(group['primary']['id'])
            
            for dup in group['duplicates']:
                dup_id = int(dup['id'])
                
                # Move Fees to Primary
                fees = self.session.query(Fee).filter_by(member_id=dup_id).all()
                for fee in fees:
                    # Check if primary already has fee for this month
                    existing = self.session.query(Fee).filter_by(member_id=primary_id, month=fee.month).first()
                    if not existing:
                        fee.member_id = primary_id
                    else:
                        # Conflict: Delete duplicate fee
                        self.session.delete(fee)
                
                # Move Attendance to Primary
                attendance = self.session.query(Attendance).filter_by(member_id=dup_id).all()
                for att in attendance:
                    att.member_id = primary_id
                    
                # Delete Duplicate Member
                dup_member = self.session.query(Member).get(dup_id)
                self.session.delete(dup_member)
                merged_count += 1
                
        self.session.commit()
        return merged_count

    def import_json_data(self, data):
        """Import legacy JSON data into database"""
        if self.legacy:
            return False, "Already in legacy mode"

        try:
            # Track ID mapping (Old String ID -> New DB ID)
            id_map = {}
            imported_members = 0
            
            # 1. Import Members
            members = data.get('members', {})
            for idx, (old_id, m_data) in enumerate(members.items()):
                # Check if already exists (by phone)
                existing = self.session.query(Member).filter_by(
                    gym_id=self.gym.id, 
                    phone=m_data.get('phone')
                ).first()
                
                if existing:
                    id_map[old_id] = existing.id
                    continue
                    
                # Create NEW member
                joined = datetime.now().date()
                if m_data.get('joined_date'):
                    try:
                        joined = datetime.strptime(m_data['joined_date'], '%Y-%m-%d').date()
                    except: pass
                    
                member = Member(
                    gym_id=self.gym.id,
                    name=m_data.get('name'),
                    phone=m_data.get('phone'),
                    email=m_data.get('email'),
                    photo_url=m_data.get('photo'),
                    joined_date=joined,
                    membership_type=m_data.get('membership_type', 'Gym'),
                    is_trial=m_data.get('is_trial', False),
                    is_active=True
                )
                self.session.add(member)
                self.session.flush() # Get new ID
                id_map[old_id] = member.id
                imported_members += 1
                
                # Commit every 50 records for Railway
                if (idx + 1) % 50 == 0:
                    self.session.commit()
                    print(f"✓ Committed {idx + 1} members")
            
            # CRITICAL: Commit remaining members
            if imported_members % 50 != 0:
                self.session.commit()
                print(f"✓ Committed final {imported_members} members")
                
            # 2. Import Fees
            fees_data = data.get('fees', {})
            for old_id, member_fees in fees_data.items():
                if old_id not in id_map: continue
                new_id = id_map[old_id]
                
                for month, info in member_fees.items():
                    # Check duplicate
                    exists = self.session.query(Fee).filter_by(member_id=new_id, month=month).first()
                    if exists: continue
                    
                    p_date = datetime.now()
                    if info.get('date'):
                        try:
                            p_date = datetime.strptime(info['date'], '%Y-%m-%d %H:%M:%S')
                        except: pass
                        
                    fee = Fee(
                        member_id=new_id,
                        month=month,
                        amount=float(info.get('amount', 0)),
                        paid_date=p_date
                    )
                    self.session.add(fee)
            
            # 3. Import Attendance
            att_data = data.get('attendance', {})
            for old_id, logs in att_data.items():
                if old_id not in id_map: continue
                new_id = id_map[old_id]
                
                for log in logs:
                    ts = datetime.now()
                    if log.get('timestamp'):
                        try:
                            ts = datetime.strptime(log['timestamp'], '%Y-%m-%d %H:%M:%S')
                        except: pass
                    
                    # Check duplicate (approximate)
                    exists = self.session.query(Attendance).filter(
                        Attendance.member_id == new_id,
                        Attendance.check_in_time == ts
                    ).first()
                    
                    if exists: continue
                    
                    att = Attendance(
                        member_id=new_id,
                        check_in_time=ts,
                        emotion=log.get('emotion'),
                        confidence=log.get('confidence')
                    )
                    self.session.add(att)
            
            self.session.commit()
            return True, f"Imported {imported_members} members successfully"
            
        except Exception as e:
            self.session.rollback()
            return False, str(e)

    def __del__(self):
        """Close session"""
        if hasattr(self, 'session'):
            self.session.close()
    
    # ==================== PAYMENT METHODS ====================

    def get_member_fees(self, member_id):
        """Get all fee records for a member (Alias for compatibility)"""
        return self.get_payment_history(member_id)

    def get_payment_history(self, member_id):
        """Get payment history for a specific member"""
        if self.legacy:
            return []
            
        try:
            fees = self.session.query(Fee).filter_by(
                member_id=int(member_id)
            ).order_by(Fee.paid_date.desc()).all()
            
            return [{
                'id': f.id,
                'member_id': f.member_id,
                'month': f.month,
                'amount': float(f.amount) if f.amount else 0.0,
                'paid_date': f.paid_date.strftime('%Y-%m-%d') if f.paid_date else None,
                'notes': f.notes or ''
            } for f in fees]
        except Exception as e:
            print(f"Error getting payment history: {str(e)}")
            return []

    def get_classes(self):
        """Get all scheduled classes - stub method"""
        # TODO: Implement class scheduling feature
        return []
    
    def add_class(self, name, day, time, instructor, capacity):
        """Add a scheduled class - stub method"""
        # TODO: Implement class scheduling feature
        return True
    def get_dashboard_stats(self, months=6):
        """
        Get ALL dashboard stats in optimized queries.
        Returns:
            - stats: Dict with counts and revenue
            - alerts: Dict with lists for UI alerts
            - charts: Dict with data for charts
        """
        if self.legacy:
            return {}, {}, {}

        today = datetime.now().date()
        current_month_str = datetime.now().strftime('%Y-%m')
        last_month_str = (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
        
        # 1. MEMBERS & REVENUE STATUS (Few Queries)
        # Get all active members
        members_query = self.session.query(Member).filter(
            Member.gym_id == self.gym.id,
            Member.is_active == True
        ).all()
        
        total_members = len(members_query)
        
        # Get payments for current and last month
        payments_query = self.session.query(Fee).filter(
            Fee.gym_id == self.gym.id,
            Fee.month.in_([current_month_str, last_month_str])
        ).all()
        
        # Process payments in memory (fast for <10k records)
        paid_member_ids = set()
        current_revenue = 0.0
        last_month_revenue = 0.0
        
        for p in payments_query:
            if p.month == current_month_str:
                current_revenue += float(p.amount)
                paid_member_ids.add(p.member_id)
            elif p.month == last_month_str:
                last_month_revenue += float(p.amount)
                
        paid_count = len(paid_member_ids)
        unpaid_count = total_members - paid_count
        
        # Calculate revenue change
        revenue_change = 0
        if last_month_revenue > 0:
            revenue_change = round(((current_revenue - last_month_revenue) / last_month_revenue) * 100, 1)

        # 2. LISTS & ALERTS
        # We need full lists for the dashboard tables, not just counts
        paid_list = []
        unpaid_list = []
        
        # Alerts (subsets)
        expiring_trials = []
        birthdays_today = []
        
        for m in members_query:
            # Create member dict for UI
            m_dict = {
                'id': m.id, 
                'name': m.name, 
                'photo_url': m.photo_url, 
                'phone': m.phone,
                'email': m.email,
                'photo': m.photo_url # Template uses 'photo' sometimes
            }
            
            # Categorize Paid/Unpaid
            if m.id in paid_member_ids:
                m_dict['status'] = 'paid'
                paid_list.append(m_dict)
            else:
                m_dict['status'] = 'unpaid'
                unpaid_list.append(m_dict)
            
            # Expiring Trials
            if m.is_trial and m.trial_end_date:
                days_left = (m.trial_end_date - today).days
                if 0 <= days_left <= 3:
                    expiring_trials.append({
                        'id': m.id, 'name': m.name, 'photo_url': m.photo_url, 
                        'days_left': days_left, 'trial_end': m.trial_end_date
                    })
            
            # Birthdays
            if m.birthday:
                if m.birthday.month == today.month and m.birthday.day == today.day:
                    birthdays_today.append({'id': m.id, 'name': m.name, 'photo_url': m.photo_url})
        
        # Limit alerts size for specific alert widgets (top 5)
        unpaid_alert = unpaid_list[:5]

        # 3. INACTIVE MEMBERS (Optimized Query)
        cutoff_date = datetime.now() - timedelta(days=14)
        
        # Subquery for max date
        subquery = self.session.query(
            Attendance.member_id,
            func.max(Attendance.created_at).label('last_checkin')
        ).group_by(Attendance.member_id).subquery()
        
        # Join
        inactive_query = self.session.query(Member, subquery.c.last_checkin)\
            .outerjoin(subquery, Member.id == subquery.c.member_id)\
            .filter(Member.gym_id == self.gym.id, Member.is_active == True)
            
        inactive_list = []
        for member, last_checkin in inactive_query.all():
            days_inactive = 999
            if last_checkin:
                days_inactive = (datetime.now() - last_checkin).days
            
            if days_inactive > 14:
                inactive_list.append({
                    'id': member.id, 'name': member.name, 'photo_url': member.photo_url,
                    'days_inactive': days_inactive
                })
        
        inactive_list.sort(key=lambda x: x['days_inactive'], reverse=True)
        inactive_list = inactive_list[:5]

        # 4. REVENUE TREND CHART (Optimized Query)
        start_date = (datetime.now().replace(day=1) - timedelta(days=30*months)).strftime('%Y-%m')
        revenue_hist = self.session.query(Fee.month, func.sum(Fee.amount))\
            .filter(Fee.gym_id == self.gym.id, Fee.month >= start_date)\
            .group_by(Fee.month).all()
            
        revenue_map = {r[0]: float(r[1]) for r in revenue_hist}
        
        revenue_trend = []
        current_dt = datetime.now()
        for i in range(months-1, -1, -1):
            year = current_dt.year
            month = current_dt.month - i
            while month <= 0: month += 12; year -= 1
            
            month_str = f"{year}-{month:02d}"
            label = datetime(year, month, 1).strftime('%b')
            revenue_trend.append({'month': label, 'revenue': revenue_map.get(month_str, 0)})

        return {
            'total_members': total_members,
            'paid_count': paid_count,
            'unpaid_count': unpaid_count,
            'revenue': current_revenue,
            'revenue_change': revenue_change,
            'expiring_count': len(expiring_trials),
            'paid_list': paid_list,
            'unpaid_list': unpaid_list
        }, {
            'unpaid': unpaid_alert,
            'expiring': expiring_trials,
            'birthdays': birthdays_today,
            'inactive': inactive_list
        }, {
            'revenue_trend': revenue_trend
        }
